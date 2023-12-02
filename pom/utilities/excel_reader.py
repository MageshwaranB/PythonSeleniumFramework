import openpyxl


def load_xl(filepath: str, sheet_name: str):
    global workbook
    global sheet
    workbook=openpyxl.load_workbook(filepath)
    sheet=workbook[sheet_name]


def get_cell_value(row_num, cell_num):
    return sheet.cell(row_num,cell_num).value

def get_data_as_list_of_tuples():
    sheet_cells=[]
    for i in range(1, sheet.max_row):
        rows=[]
        for cell in sheet[i+1]:
            rows.append(cell.value)
        sheet_cells.append(tuple(rows))
    return sheet_cells
